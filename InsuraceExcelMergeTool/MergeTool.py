from openpyxl import load_workbook
from openpyxl import workbook
import os
import datetime
import shutil

excel_name = os.listdir('./待合并文件')
filepath = os.getcwd()
wbw = load_workbook(filepath + '\\已合并文件\\请复制别剪切.xlsx')
sameFile = workbook.Workbook()
sameSheet = sameFile.worksheets[0]
for en in excel_name:
    enn = filepath + '\\待合并文件\\' + en
    wbr = load_workbook(enn)
    sheetIndex = -1
    for sheetN in wbr.sheetnames:
        sheetIndex += 1
        wbwSheet = wbw.worksheets[sheetIndex]
        wbrSheet = wbr[sheetN]
        rowNum = 0
        # 去除最后一行的汇总
        for count in range(wbrSheet.max_row):
            rowNum += 1
            if rowNum <= 2:
                cellI = 1
                for cell in wbrSheet[rowNum]:
                    wbwSheet.cell(rowNum, cellI).value = cell.value
                    cellI += 1
            else:
                # 如果第一个单元格为空则跳过，因为最后一行是汇总数
                if wbrSheet.cell(rowNum, 1).value is None:
                    continue
                # 判断是否是相同行
                isSame = False
                sameRow = 0
                for row in wbwSheet.rows:
                    sameRow += 1
                    pd = 0
                    isSame = True
                    for cell in row:
                        pd += 1
                        if wbrSheet.cell(rowNum, pd).value != cell.value:
                            isSame = False
                            break
                    # if wbrSheet.cell(rowNum, 1).value != row[0].value:
                    #     isSame = False
                    # elif wbrSheet.cell(rowNum, 19).value != row[18].value:
                    #     isSame = False
                    # elif wbrSheet.cell(rowNum, 20).value != row[19].value:
                    #     isSame = False
                    # elif wbrSheet.cell(rowNum, 45).value != row[44].value:
                    #     isSame = False

                    if isSame:
                        sameSheet.cell(1, 1).value = '待合并文件为'
                        sameSheet.cell(1, 2).value = '待合并文件相同的行数为'
                        sameSheet.cell(1, 3).value = '合并结果与其相同的行数为'
                        sameSheet.cell(sameSheet.max_row + 1, 1).value = en
                        sameSheet.cell(sameSheet.max_row, 2).value = rowNum
                        sameSheet.cell(sameSheet.max_row, 3).value = sameRow
                        sameFile.save(filepath + '\\相同行\\' + datetime.datetime.now().strftime('%Y-%m-%d') + '.xlsx')
                        break
                if not isSame:
                    maxRow = wbwSheet.max_row
                    cellI = 1
                    for cell in wbrSheet[rowNum]:
                        wbwSheet.cell(maxRow + 1, cellI).value = cell.value
                        cellI += 1
    shutil.move(enn, filepath + '\\合并完自动转入\\' + en)

wbw.save(filepath + '\\已合并文件\\请复制别剪切.xlsx')
